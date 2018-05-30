import os
from os.path import isfile, join
import time
import openpyxl

from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.common.keys import Keys

driver = webdriver.Firefox()
# driver = webdriver.Firefox(executable_path='C:\\PROJECTS\\uni-selenium\\geckodriver.exe')
driver.implicitly_wait(1)


def wait_by_id(el_id, seconds=10):
    element = WebDriverWait(driver, seconds).until(
        EC.presence_of_element_located((By.ID, el_id))
    )
    return element


def wait_by_classname(el_class, seconds=10):
    element = WebDriverWait(driver, seconds).until(
        EC.presence_of_element_located((By.CLASS_NAME, el_class))
    )
    return element


def wait_by_css(selector, seconds=10):
    elements = WebDriverWait(driver, seconds).until(
        EC.presence_of_element_located((By.CLASS_NAME, selector))
    )
    return elements


def wait_by_name(selector, seconds=10):
    elements = WebDriverWait(driver, seconds).until(
        EC.presence_of_element_located((By.NAME, selector))
    )
    return elements


def get_person_list():
    files_folder_name = "input"
    for file in os.listdir(files_folder_name):
        if file.endswith(".xlsx"):
            wb = openpyxl.load_workbook(join(files_folder_name, file))
            ws = wb.get_sheet_by_name(wb.get_sheet_names()[0])
            result = []
            for row in ws.iter_rows():
                current_row = row[0].row
                person = {
                    "sex": ws["A" + str(current_row)].value,
                    "first_name": ws["B" + str(current_row)].value,
                    "second_name": ws["C" + str(current_row)].value,
                    "last_name": ws["D" + str(current_row)].value,
                    "organization": ws["E" + str(current_row)].value,
                    "position": ws["F" + str(current_row)].value
                }
                if None not in person.values():
                    if current_row > 1:
                        result.append(person)
                        print(person)
                else:
                    print("Error: 'None' in field: ", person)


driver.get("https://uni.urfu.ru/fx/")
wait_by_id('login', 10).send_keys('n.v.ignatchenko@urfu.ru')
wait_by_id('password', 10).send_keys('Ye;yj,jkmitrjlf')
wait_by_name('LogonFormSubmit', 10).click()
# driver.execute_script('document.querySelector(".menuitem_submain_ul").style.display = "block"')
# driver.find_element_by_xpath('//*[@id="pohour_leaner_groups_mi_link"]').click()
driver.get(
    "https://uni.urfu.ru/fx/uni/ru.naumen.uni.published_jsp?uuid=fakeobUNI_HourpayedCoreRoot&activeComponent=ActiveGroups")
wait_by_id('ActiveGroups.LearnerGroups.LearnerGroup', 10).click()
driver.switch_to.window(driver.window_handles[1])
WebDriverWait(driver, 10).until(lambda d: d.title != "")
wait_by_id('title', 10).send_keys('номер групппы')
time.sleep(2)
# wait_by_id('cancel',20).click().perform()
wait_by_id('cancel', 20).click()
driver.switch_to.window(driver.window_handles[0])
WebDriverWait(driver, 10).until(lambda d: d.title != "")
wait_by_id('TitleFilter').send_keys('ДК-2017м')
wait_by_id('TitleFilter').send_keys(Keys.ENTER)
wait_by_id('ActiveGroups.LearnerGroups', 20).find_elements_by_tag_name('a')[0].click()

# group = lambda: WebDriverWait(driver, 1).until(
#                 EC.presence_of_element_located((By.CSS_SELECTOR, '[id*="ActiveGroups"] a'))
#             )
# group().click()

driver.execute_script('document.querySelector(\'[id*="ActiveGroups"] tr:nth-child(2) a\').click()')
wait_by_id('LearnerGroupTab.Learners.AddLearnerShort').click()
driver.switch_to.window(driver.window_handles[1])
WebDriverWait(driver, 10).until(lambda d: d.title != "")
wait_by_id('lastName', 10).send_keys('Фамилия')
wait_by_id('firstName', 10).send_keys('Имя')
wait_by_id('middleName', 10).send_keys('Имя')
# wait_by_id('BirthDate', 10).send_keys('08.05.2018')
sex = 'м'
if sex == "ж":
    wait_by_id('Sex_pstcim18gg3ig0000glddc6svsi80ce8', 10).click()
if sex == "м":
    wait_by_id('Sex_pstcim18gg3ig0000glddc6svsi80ce4', 10).click()
wait_by_id('placeOfEmployment', 10).send_keys('Место работы')
wait_by_id('post', 10).send_keys('Должность')
